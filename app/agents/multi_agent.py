from typing import TypedDict, Literal
from langgraph.graph import StateGraph, END
from langgraph.checkpoint.sqlite import SqliteSaver
import logging
import sqlite3
from openpyxl import load_workbook
from pathlib import Path
import json, os, datetime
#添加rag
from app.rag.vector_store import VectorStore

BASE_DIR = Path(__file__).parent.parent.parent
CHROMA_DIR = BASE_DIR / "chroma_db"

vector_store = VectorStore(persist_dir=str(CHROMA_DIR), collection_name="faq")

base_path = Path(__file__).parent.parent.parent
sample_path = base_path / "sample.xlsx"

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.info(f"VectorStore 使用的集合: {vector_store.collection_name}")
logging.info(f"VectorStore 使用的持久化目录: {vector_store.persist_dir}")


class AgentState(TypedDict):
    user_request: str
    task_type: str
    file_path: str          # 新增：Excel 文件路径
    parsed_data: list[dict]    # 新增：解析后的数据
    worker_result: str
    final_output: str
    generated_json: str          # 新增：存储 JSON 字符串
    steps: list[str]          # 步骤列表，如 ["解析Excel", "校验数据", "生成报告"]
    current_step_index: int   # 当前执行的步骤索引（从 0 开始）


def supervisor_node(state: AgentState) -> dict:
    request = state["user_request"]
    if "解析" in request:
        task_type = "解析"
    elif "校验" in request:
        task_type = "校验"
    elif "检索" in request or "查" in request:#关键词触发rag
        task_type = "rag"
    else:
        task_type = "生成"

    parts = request.split()
    file_path = parts[-1] if len(parts) > 1 else str(sample_path)    

    return {
        "task_type": task_type,
        "file_path": file_path
    }

# def plan_node(state: AgentState) -> dict:
#     """根据用户请求生成执行步骤列表"""
#     request = state.get("user_request", "")
#         # 简单规则：根据关键词生成步骤
#     if "解析" in request and "校验" in request and "生成" in request:
#         steps = ["解析Excel", "校验数据", "生成报告"]
#     elif "解析" in request and "校验" in request:
#         steps = ["解析Excel", "校验数据"]
#     elif "解析" in request:
#         steps = ["解析Excel"]
#     elif "校验" in request:
#         # 假设之前已经解析过了，直接校验
#         steps = ["校验数据"]
#     elif "生成" in request:
#         steps = ["生成报告"]
#     else:
#         steps = ["解析Excel", "校验数据", "生成报告"]  # 默认全流程
#     return {
#         "steps": steps,
#         "current_step_index": 0,
#         "task_type": "plan_execute"  # 标记当前模式
#     }

# def executor_node(state: AgentState) -> dict:
#     """执行当前步骤，如果还有下一步则继续"""
#     steps = state.get("steps",[])
#     idx = state.get("current_step_index", 0)

#     # 检查是否所有步骤都已完成
#     if idx >= len(steps):
#         return {"final_output": "所有步骤执行完成"}
    
#     current_step = steps[idx]
#     logging.info("执行步骤 %d/%d: %s", idx + 1, len(steps), current_step)

#     # 根据步骤名称执行对应逻辑
#     result = ""
#     if current_step == "解析Excel": 
#         # 调用 parse_worker 的逻辑
#         file_path = state.get("file_path", "")
#         if os.path.exists(file_path):
#             wb = load_workbook(file_path, data_only=True)
#             ws = wb.active
#             headers = [cell.value for cell in ws[1]]
#             data = []
#             for row in ws.iter_rows(min_row=2, values_only=True):
#                 if any(cell is not None for cell in row):
#                     data.append(dict(zip(headers, row)))
#             result = f"解析成功：读取了 {len(data)} 行数据，表头：{headers}"
#         else:
#             result = f"错误：文件 {file_path} 不存在"

#     elif current_step == "校验数据":
#         # 调用 validate_worker 的逻辑
#         data = state.get("parsed_data", [])
#         if not data:
#             result = "校验失败：没有数据"
#         else:
#             # 简化校验（实际可用 validate_worker 的逻辑）
#             required = {"姓名", "年龄"}
#             available = set(data[0].keys()) if data else set()
#             missing = required - available
#             if missing:
#                 result = f"校验失败：缺少列 {missing}"
#             else:
#                 result = f"校验通过：{len(data)} 行数据"
#     elif current_step == "生成报告":
#         # 调用 generate_worker 的逻辑
#         data = state.get("parsed_data", [])
#         if not data:
#             result = "生成失败：没有数据"
#         else:
#             output = {
#                 "total_rows": len(data),
#                 "columns": list(data[0].keys()) if data else [],
#                 "data": data
#             }
#             json_str = json.dumps(output, ensure_ascii=False, indent=2)
#             result = f"生成成功：已生成 {len(data)} 行数据的 JSON"
#     else:
#         result = f"未知步骤：{current_step}"
    
#     # 更新状态：记录当前步骤结果，索引+1
#     return {
#         "worker_result": result,
#         "current_step_index": idx + 1
#     }        





#定义方法的返回值为上述的字典类型
# def parse_worker(state: AgentState) -> dict:
# #     """解析 Worker"""
# #     # raise RuntimeError("模拟错误：Excel 解析失败，文件格式不正确")
#     return {"worker_result": f"解析了: {state['user_request']}"}

def parse_worker(state: AgentState) -> dict:
    file_path = state.get("file_path", "sample.xlsx")
    # raise RuntimeError(f"模拟业务错误：Excel 文件 {file_path} 缺少 '订单号' 列")
    if not file_path.endswith(".xlsx"):
        raise ValueError("文件格式不正确，请提供一个 Excel 文件 (.xlsx)")
    
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        # 读取数据行
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))

        result = f"解析成功：读取了 {len(data)} 行数据，表头：{headers}"

        return {
            "worker_result": result,
            "parsed_data": data   # 可选的额外字段
        }    

    except Exception as e:
        return {"worker_result": f"解析失败：{str(e)}"}


def validate_worker(state: AgentState) -> dict:
    # """校验 Worker"""
    # return {"worker_result": f"校验了: {state['user_request']}"}
    data = state.get("parsed_data", [])
    if not data:
        return {"worker_result": "校验失败：没有解析到数据"}
    
    required_columns = ["订单号", "客户名称", "金额"]
    first_row = data[0]
    available_columns = set(first_row.keys())
    missing=required_columns - available_columns

    errors = []
    for idx,row in enumerate(data, start=2):  # 从第2行开始，因为第1行是表头
        age = row.get("年龄")
        if not isinstance(age, (int, float)) or age <= 0:
            errors.append(f"第 {idx} 行年龄无效")
    if errors:
        return {"worker_result": f"校验失败：发现 {len(errors)} 条数据异常"}

    return {"worker_result": f"校验通过：{len(data)} 行数据全部有效"}    


def generate_worker(state: AgentState) -> dict:
    # """生成 Worker"""
    # return {"worker_result": f"生成了: {state['user_request']}"}
    """将校验通过的数据生成 JSON 格式输出"""
    data = state.get("parsed_data", [])

    if not data:
        return {"worker_result": "生成失败：没有可用数据"}
    
    output = {
        "total_rows": len(data),
        "columns": list(data[0].keys()) if data else [],
        "data": data,
        "summary": {
            "generated_at": str(datetime.now()),
            "version": "1.0"
        }
    }

    json_str = json.dumps(output, ensure_ascii=False, indent=2)

    logging.info("生成 JSON 成功，共 %d 行数据", len(data))
    logging.debug("JSON 预览: %s", json_str[:200] + "..." if len(json_str) > 200 else json_str)

    return {
        "worker_result": f"生成成功：已生成 {len(data)} 行数据的 JSON",
        "generated_json": json_str  # 可选：存入 State，供后续节点使用
    }

def final_aggregator(state: AgentState) -> dict:
    """汇总结果"""
    return {"final_output": f"【完成】{state['worker_result']}"}


logging.info(f"📂 当前持久化目录: {os.path.abspath(vector_store.persist_dir)}")
def rag_worker(state: AgentState) -> dict:
    # query = state.get("user_request", "")
    # logging.info(f"🔍 rag_worker 收到的 query: {query}")  # 新增
    # docs = vector_store.retrieve(query, top_k=2)#引用rag的retrieve方法
    # logging.info(f"📊 检索到的文档数: {len(docs)}")        # 新增
    # if docs:
    #     context = "\n".join([d["document"] for d in docs])
    #     return {"worker_result": f"✅ 检索到 {len(docs)} 条相关资料：\n{context}"}
    # return {"worker_result": "⚠️ 未检索到相关资料"}
    query = state.get("user_request", "")
    logging.info(f"🔍 rag_worker 收到的 query: {query}")
    logging.info(f"📊 向量库文档总数: {vector_store.count()}")   # 新增
    docs = vector_store.retrieve(query, top_k=2)
    logging.info(f"📊 检索到的文档数: {len(docs)}")
    if docs:
        context = "\n".join([d["document"] for d in docs])
        return {"worker_result": f"✅ 检索到 {len(docs)} 条相关资料：\n{context}"}
    return {"worker_result": "⚠️ 未检索到相关资料"}


#定义返回值为三个方法的枚举类型
def route_after_supervisor(state: AgentState) -> Literal["parse_worker", "validate_worker", "generate_worker", "rag_worker"]:
    task = state.get("task_type")
    if task == "解析":
        return "parse_worker"
    elif task == "校验":
        return "validate_worker"
    elif task == "rag":
        return "rag_worker"#新增rag分支
    else:
        return "generate_worker"
    
def route_after_excutor(state: AgentState) -> Literal["executor_node", "aggregator"]:
    """判断 Executor 是否还有下一步"""
    steps = state.get("steps", [])
    idx = state.get("current_step_index", 0)
    if idx < len(steps):
        return "executor_node"   # 继续执行下一步
    else:
        return "aggregator"      # 所有步骤完成，去汇总
    
builder = StateGraph(AgentState)
builder.add_node("supervisor_node", supervisor_node)
builder.add_node("parse_worker", parse_worker)
builder.add_node("validate_worker", validate_worker)
builder.add_node("rag_worker", rag_worker)
builder.add_node("generate_worker", generate_worker)
builder.add_node("final_aggregator", final_aggregator)

builder.set_entry_point("supervisor_node")
builder.add_conditional_edges("supervisor_node", route_after_supervisor)


builder.add_edge("rag_worker", "final_aggregator")
builder.add_edge("parse_worker", "final_aggregator")
builder.add_edge("validate_worker", "final_aggregator")
builder.add_edge("generate_worker", "final_aggregator")
builder.add_edge("final_aggregator", END)

#引入轻量数据库 
conn = sqlite3.connect("checkpoints.db", check_same_thread=False)
# ========== 集成 Checkpoint ==========
logging.info("集成 Checkpoint: SqliteSaver")

memory = SqliteSaver(conn)


print("所有节点:", builder.nodes)
print("所有边:", builder.edges)
graph = builder.compile(checkpointer=memory)
print(graph)

if __name__ == "__main__":
    config = {"configurable": {"thread_id": "session_001"}}

    logging.info("🔄 尝试从 Checkpoint 恢复...")
    resumed = graph.invoke(None, config=config)
    logging.info("✅ 恢复后的结果: %s", resumed.get("final_output"))

    # result1 = graph.invoke({"user_request": "请解析这段文本"})
    # 第一次执行
    try:
        result = graph.invoke(
            {"user_request": "请解析这份Excel"},
            config=config
        )
        logging.info("测试1: %s", result["final_output"])
    except Exception as e:
        logging.error("执行过程中发生错误: %s", str(e))    
    
    # 第二次执行（同一个 thread_id，会从上次的 State 继续）
    # result2 = graph.invoke(
    #     {"user_request": "校验数据"},
    #     config=config
    # )
    # logging.info("测试2: %s", result2["final_output"])
    # print("测试1:", result1["final_output"])

    # history = graph.get_state_history(config)
    # for state_snapshot in history:
    #     print(f"🔹 节点: {state_snapshot}")
    #     print(f"🔹 节点: {state_snapshot.values}, 状态: {state_snapshot.tasks}")

    # 获取当前状态（最新snapshot）
    current = graph.get_state(config)
    logging.info("当前状态 next: %s", current.next)

    resumed_result = graph.invoke(None, config=config)
    logging.info("恢复后的结果: %s", resumed_result)

    config = {"configurable": {"thread_id": "rag_test"}}

    # ===== 新增：测试 RAG 检索 =====
    logging.info("="*40)
    logging.info("测试 RAG 检索...")
    rag_result = graph.invoke(
        {"user_request": "检索一下美国仓发货时效"},
        config=config
    )
    logging.info("RAG 检索结果: %s", rag_result.get("final_output"))